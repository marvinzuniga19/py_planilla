import os
from datetime import datetime
from pathlib import Path

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExportUtils:
    @staticmethod
    def export_to_pdf(data, filename):
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab no está instalado")
        
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20
        )
        
        elements.append(Paragraph("Planilla de Nómina", title_style))
        elements.append(Paragraph(f"Período: {data[0]['month']}/{data[0]['year']}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        headers = ['Nombre', 'DPI', 'Cargo', 'Salario Base', 'Horas Extras', 
                   'INATEC', 'INSS', 'Otros', 'Salario Neto']
        table_data = [headers]
        
        for row in data:
            table_data.append([
                row.get('name', ''),
                row.get('dpi', ''),
                row.get('position', ''),
                f"${row.get('base_salary', 0):.2f}",
                f"${row.get('extra_hours_amount', 0):.2f}",
                f"${row.get('total_afp', 0):.2f}",
                f"${row.get('total_isss', 0):.2f}",
                f"${row.get('other_discounts', 0):.2f}",
                f"${row.get('net_salary', 0):.2f}"
            ])
        
        table = Table(table_data, colWidths=[80, 80, 60, 60, 60, 50, 50, 50, 60])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        return filename

    @staticmethod
    def export_to_excel(data, filename):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nómina"
        
        headers = ['Nombre', 'DPI', 'Cargo', 'Salario Base', 'Horas Extras', 
                   'INATEC', 'INSS', 'Otros Descuentos', 'Salario Neto']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, row_data in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=row_data.get('name', ''))
            ws.cell(row=row_idx, column=2, value=row_data.get('dpi', ''))
            ws.cell(row=row_idx, column=3, value=row_data.get('position', ''))
            ws.cell(row=row_idx, column=4, value=row_data.get('base_salary', 0))
            ws.cell(row=row_idx, column=5, value=row_data.get('extra_hours_amount', 0))
            ws.cell(row=row_idx, column=6, value=row_data.get('total_afp', 0))
            ws.cell(row=row_idx, column=7, value=row_data.get('total_isss', 0))
            ws.cell(row=row_idx, column=8, value=row_data.get('other_discounts', 0))
            ws.cell(row=row_idx, column=9, value=row_data.get('net_salary', 0))
        
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        wb.save(filename)
        return filename
